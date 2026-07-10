from flask import (Flask, render_template_string, request, redirect,
                   url_for, session, send_from_directory, abort)
from werkzeug.security import generate_password_hash, check_password_hash
from jinja2 import DictLoader
import sqlite3, os, json, secrets, smtplib, io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'users.db')
EBOOKS_DIR = BASE_DIR

with open(os.path.join(BASE_DIR, 'ebooks_metadata.json'), 'r', encoding='utf-8') as _f:
    EBOOKS = json.load(_f)

_CAT_ORDER = [
    'Clássicos Brasileiros', 'Classicos Brasileiros',
    'Livros em Inglês',
    'IA e Tecnologia', 'Negócios', 'Finanças', 'Importação',
    'Liderança', 'Desenvolvimento Pessoal', 'Varejo', 'Saúde e Gestão',
    'Automóveis',
    'Culinária', 'Educação Infantil', 'Estilo de vida', 'Idiomas', 'Saúde',
]
def _cat_rank(cat):
    cl = cat.lower()
    for i, o in enumerate(_CAT_ORDER):
        if o.lower() in cl or cl in o.lower():
            return i
    return 99

CATEGORIES = {}
for _eb in EBOOKS:
    CATEGORIES.setdefault(_eb['categoria'], []).append(_eb)
CATEGORIES = dict(sorted(CATEGORIES.items(), key=lambda kv: _cat_rank(kv[0])))

CAT_ICONS = {
    'IA e Tecnologia': 'bi-cpu-fill',
    'Negócios': 'bi-briefcase-fill',
    'Finanças': 'bi-cash-coin',
    'Importação': 'bi-globe-americas',
    'Liderança': 'bi-trophy-fill',
    'Desenvolvimento Pessoal':'bi-lightbulb-fill',
    'Varejo': 'bi-bag-fill',
    'Saúde e Gestão': 'bi-heart-pulse-fill',
    'Livros em Inglês': 'bi-globe2',
    'Clássicos Brasileiros': 'bi-book-fill',
    'Automóveis': 'bi-car-front-fill',
    'Culinária': 'bi-fire',
    'Educação Infantil': 'bi-pencil-fill',
    'Estilo de vida': 'bi-stars',
    'Idiomas': 'bi-translate',
    'Saúde': 'bi-heart-pulse-fill',
}
CAT_COLORS = {
    'IA e Tecnologia': '#00BCD4',
    'Negócios': '#F39C12',
    'Finanças': '#27AE60',
    'Importação': '#FF6B35',
    'Liderança': '#C0392B',
    'Desenvolvimento Pessoal':'#9B59B6',
    'Varejo': '#E91E63',
    'Saúde e Gestão': '#16A085',
    'Livros em Inglês': '#2196F3',
    'Clássicos Brasileiros': '#4CAF50',
    'Automóveis': '#E74C3C',
    'Culinária': '#FF9800',
    'Educação Infantil': '#9C27B0',
    'Estilo de vida': '#FF6B9D',
    'Idiomas': '#009688',
    'Saúde': '#1ABC9C',
}

def _cat_icon(cat):
    for k, v in CAT_ICONS.items():
        if k.lower() in cat.lower() or cat.lower() in k.lower():
            return v
    return 'bi-folder-fill'

def _cat_color(cat):
    for k, v in CAT_COLORS.items():
        if k.lower() in cat.lower() or cat.lower() in k.lower():
            return v
    return '#607D8B'

MAIL_USER = os.environ.get('MAIL_USERNAME', '')
MAIL_PASS = os.environ.get('MAIL_PASSWORD', '')

def _send_email(to_addr, subject, html):
    if not (MAIL_USER and MAIL_PASS and to_addr):
        return False
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = 'Clube do Saber <{}>'.format(MAIL_USER)
        msg['To'] = to_addr
        msg.attach(MIMEText(html, 'html', 'utf-8'))
        with smtplib.SMTP('smtp.gmail.com', 587) as s:
            s.starttls()
            s.login(MAIL_USER, MAIL_PASS)
            s.send_message(msg)
        return True
    except Exception as ex:
        print('[mail] {}'.format(ex))
        return False

def _welcome_html(username, password, login_url):
    return (
        '<div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">'
        '<div style="background:linear-gradient(135deg,#16213e,#533483);padding:28px 32px;border-radius:12px 12px 0 0;text-align:center">'
        '<h1 style="color:#fff;margin:0;font-size:1.4rem">&#128218; Clube do Saber</h1>'
        '</div>'
        '<div style="background:#fff;padding:28px 32px;border-radius:0 0 12px 12px;border:1px solid #eee;border-top:none">'
        '<p>Ola, <strong>{u}</strong>! Seu acesso ao Clube do Saber foi criado.</p>'
        '<div style="background:#f8f5ff;border-left:4px solid #6C3483;padding:16px;border-radius:6px;margin:16px 0">'
        '<p style="margin:0 0 8px"><strong>Usuario:</strong> {u}</p>'
        '<p style="margin:0"><strong>Senha:</strong> {p}</p>'
        '</div>'
        '<a href="{l}" style="display:inline-block;background:#6C3483;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:bold">Acessar agora</a>'
        '</div></div>'
    ).format(u=username, p=password, l=login_url)

def _reset_html(username, reset_url):
    return (
        '<div style="font-family:Arial,sans-serif;max-width:500px;margin:auto">'
        '<div style="background:linear-gradient(135deg,#16213e,#533483);padding:28px 32px;border-radius:12px 12px 0 0;text-align:center">'
        '<h1 style="color:#fff;margin:0;font-size:1.4rem">&#128218; Clube do Saber</h1>'
        '</div>'
        '<div style="background:#fff;padding:28px 32px;border-radius:0 0 12px 12px;border:1px solid #eee;border-top:none">'
        '<p>Ola, <strong>{u}</strong>! Clique abaixo para redefinir sua senha:</p>'
        '<a href="{r}" style="display:inline-block;background:#6C3483;color:#fff;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:bold">Redefinir senha</a>'
        '<p style="color:#888;font-size:.85rem;margin-top:16px">Link expira em 1 hora.</p>'
        '</div></div>'
    ).format(u=username, r=reset_url)

TMPLS = {}
TMPLS['base.html'] = '<!doctype html>\n<html lang="pt-BR">\n<head>\n<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n<title>{% block title %}Clube do Saber{% endblock %}</title>\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">\n<style>\n:root{--brand:#6C3483;--brand-dark:#4a2360;--brand-light:#9b59b6;--sidebar-bg:#16213e;--sidebar-w:240px;--nav-h:56px}\nbody{background:#f0f2f5;font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\',sans-serif}\n.top-nav{background:var(--sidebar-bg);height:var(--nav-h);position:sticky;top:0;z-index:1030;box-shadow:0 2px 8px rgba(0,0,0,.3)}\n.sidebar{width:var(--sidebar-w);min-width:var(--sidebar-w);background:#1a1a2e;height:calc(100vh - var(--nav-h));position:sticky;top:var(--nav-h);overflow-y:auto;border-right:1px solid rgba(255,255,255,.05)}\n.sidebar-header{color:rgba(255,255,255,.4);font-size:.7rem;letter-spacing:.08em;padding:.75rem 1rem .25rem}\n.sidebar-link{display:flex;align-items:center;padding:.55rem 1rem;color:rgba(255,255,255,.65);text-decoration:none;font-size:.875rem;transition:background .15s;border-left:3px solid transparent}\n.sidebar-link:hover{background:rgba(255,255,255,.07);color:#fff}\n.sidebar-link.active{background:rgba(108,52,131,.25);color:#fff;border-left-color:var(--brand-light)}\n.sidebar-link i{margin-right:.6rem}\n.main-content{flex:1;min-width:0;padding:2rem}\n.btn-brand{background:var(--brand);color:#fff;border:none}.btn-brand:hover{background:var(--brand-dark);color:#fff}\n</style>{% block head %}{% endblock %}\n</head>\n<body>\n<nav class="top-nav d-flex align-items-center px-3 gap-3">\n <a class="text-white text-decoration-none fw-bold d-flex align-items-center gap-2" href="{{ url_for(\'dashboard\') }}">\n <i class="bi bi-book-half" style="font-size:1.4rem;color:var(--brand-light)"></i>\n <span style="font-size:1.1rem">Clube do Saber</span>\n </a>\n <form class="d-none d-sm-flex" action="{{ url_for(\'search\') }}" method="get" style="flex:1;max-width:340px;margin-left:1rem"><div class="input-group input-group-sm"><input type="search" name="q" class="form-control rounded-start" placeholder="Pesquisar livros..." value="{{ request.args.get(\'q\',\'\') }}"><button class="btn btn-sm" style="background:#6C3483;color:#fff;border:none" type="submit"><i class="bi bi-search"></i></button></div></form> <div class="ms-auto d-flex align-items-center gap-3">\n <span class="text-white-50 small">{{ session.username }}</span>\n {% if session.is_admin %}<a href="{{ url_for(\'admin\') }}" class="btn btn-sm btn-outline-light"><i class="bi bi-gear"></i> Admin</a>{% endif %}\n <a href="{{ url_for(\'logout\') }}" class="btn btn-sm btn-outline-danger">Sair</a>\n </div>\n</nav>\n<div class="d-flex">\n <div class="sidebar d-none d-md-flex flex-column">\n <div class="sidebar-header">MENU</div>\n <a href="{{ url_for(\'dashboard\') }}" class="sidebar-link {% if request.endpoint==\'dashboard\' %}active{% endif %}"><i class="bi bi-house-fill"></i> Inicio</a>\n <div class="sidebar-header mt-2">CATEGORIAS</div>\n {% for cat in categories %}<a href="{{ url_for(\'dashboard\') }}#cat-{{ loop.index }}" class="sidebar-link"><i class="bi {{ cat_icon(cat) }}"></i> {{ cat }}</a>{% endfor %}\n </div>\n <div class="main-content">\n {% with messages = get_flashed_messages(with_categories=true) %}{% if messages %}{% for c,m in messages %}<div class="alert alert-{{c}} alert-dismissible fade show">{{m}}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}{% endwith %}\n {% block content %}{% endblock %}\n </div>\n</div>\n<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>\n</body></html>'

TMPLS['login.html'] = '<!doctype html><html lang="pt-BR"><head>\n<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n<title>Clube do Saber - Login</title>\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">\n<style>body{background:linear-gradient(135deg,#16213e 0%,#0f3460 50%,#533483 100%);min-height:100vh;display:flex;align-items:center;justify-content:center}.card{border:none;border-radius:16px;box-shadow:0 20px 60px rgba(0,0,0,.4)}.brand-icon{width:64px;height:64px;background:linear-gradient(135deg,#6C3483,#9b59b6);border-radius:16px;display:flex;align-items:center;justify-content:center;margin:0 auto;font-size:2rem;color:#fff}</style>\n</head><body>\n<div class="card p-4" style="width:100%;max-width:400px">\n <div class="text-center mb-4">\n <div class="brand-icon mb-3"><i class="bi bi-book-half"></i></div>\n <h4 class="fw-bold">Clube do Saber</h4>\n <p class="text-muted small mb-0">Biblioteca de Ebooks Profissionais</p>\n </div>\n {% if error %}<div class="alert alert-danger py-2">{{ error }}</div>{% endif %}\n <form method="post">\n <div class="mb-3"><label class="form-label fw-semibold">Usuario</label><input name="username" class="form-control" required autofocus></div>\n <div class="mb-4"><label class="form-label fw-semibold">Senha</label><input name="password" type="password" class="form-control" required></div>\n <button class="btn w-100 text-white fw-bold" style="background:#6C3483">Entrar</button>\n </form>\n <div class="text-center mt-3"><a href="{{ url_for(\'forgot\') }}" class="text-muted small">Esqueci minha senha</a></div>\n</div></body></html>'

TMPLS['forgot.html'] = '<!doctype html><html lang="pt-BR"><head>\n<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n<title>Clube do Saber - Recuperar Senha</title>\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">\n<style>body{background:linear-gradient(135deg,#16213e,#533483);min-height:100vh;display:flex;align-items:center;justify-content:center}.card{border:none;border-radius:16px;box-shadow:0 20px 60px rgba(0,0,0,.4)}</style>\n</head><body>\n<div class="card p-4" style="width:100%;max-width:400px">\n <h5 class="fw-bold mb-1">Recuperar senha</h5>\n <p class="text-muted small mb-3">Informe seu e-mail para receber o link.</p>\n {% if msg %}<div class="alert alert-{{ msg_type }} py-2">{{ msg }}</div>{% endif %}\n <form method="post">\n <div class="mb-3"><label class="form-label fw-semibold">E-mail</label><input name="email" type="email" class="form-control" required autofocus></div>\n <button class="btn w-100 text-white" style="background:#6C3483">Enviar link</button>\n </form>\n <div class="text-center mt-3"><a href="{{ url_for(\'login\') }}" class="text-muted small">Voltar ao login</a></div>\n</div></body></html>'

TMPLS['reset.html'] = '<!doctype html><html lang="pt-BR"><head>\n<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n<title>Clube do Saber - Nova Senha</title>\n<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">\n<style>body{background:linear-gradient(135deg,#16213e,#533483);min-height:100vh;display:flex;align-items:center;justify-content:center}.card{border:none;border-radius:16px;box-shadow:0 20px 60px rgba(0,0,0,.4)}</style>\n</head><body>\n<div class="card p-4" style="width:100%;max-width:400px">\n <h5 class="fw-bold mb-3">Criar nova senha</h5>\n {% if error %}<div class="alert alert-danger py-2">{{ error }}</div>{% endif %}\n {% if expired %}<div class="alert alert-warning">Link expirado. <a href="{{ url_for(\'forgot\') }}">Solicite outro.</a></div>\n {% else %}\n <form method="post">\n <div class="mb-3"><label class="form-label fw-semibold">Nova senha</label><input name="password" type="password" class="form-control" minlength="6" required autofocus></div>\n <div class="mb-4"><label class="form-label fw-semibold">Confirmar senha</label><input name="password2" type="password" class="form-control" minlength="6" required></div>\n <button class="btn w-100 text-white" style="background:#6C3483">Salvar senha</button>\n </form>\n {% endif %}\n</div></body></html>'

TMPLS['dashboard.html'] = '{% extends "base.html" %}\n{% block title %}Biblioteca - Clube do Saber{% endblock %}\n{% block head %}<style>\n.cat-header{border-bottom:3px solid;padding-bottom:.4rem;margin-bottom:1.2rem;font-size:1.05rem;font-weight:700}\n.book-card{background:#fff;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 2px 8px rgba(0,0,0,.07);display:flex;flex-direction:column;height:100%;border-top:4px solid var(--cc);transition:box-shadow .15s}\n.book-card:hover{box-shadow:0 6px 18px rgba(0,0,0,.13)}\n.book-title{font-weight:700;font-size:.95rem;line-height:1.3;color:#1a1a2e;margin-bottom:.3rem}\n.book-sub{font-size:.8rem;color:#6c757d;line-height:1.3;flex:1}\n.btn-dl{font-size:.8rem;padding:.35rem .9rem;border-radius:8px;margin-top:.9rem}\n</style>{% endblock %}\n{% block content %}\n<h4 class="fw-bold mb-4" style="color:#1a1a2e"><i class="bi bi-book-half me-2" style="color:#6C3483"></i>Biblioteca</h4>\n{% for cat, ebooks in categories.items() %}\n<section id="cat-{{ loop.index }}" class="mb-5">\n <div class="cat-header d-flex align-items-center gap-2" style="border-color:{{ cat_color(cat) }};color:{{ cat_color(cat) }}">\n <i class="bi {{ cat_icon(cat) }}"></i> {{ cat }}\n <span class="badge ms-1 fw-normal" style="background:{{ cat_color(cat) }};font-size:.7rem">{{ ebooks|length }}</span>\n </div>\n <div class="row g-3">\n {% for book in ebooks %}\n <div class="col-6 col-md-4 col-lg-3">\n <div class="book-card" style="--cc:{{ cat_color(book.categoria) }}">\n <div class="book-title">{{ book.titulo }}</div>\n <div class="book-sub">{{ book.subtitulo }}</div>\n <a href="{{ url_for(\'download\', filename=book.filename) }}" class="btn btn-sm btn-dl text-white" style="background:{{ cat_color(book.categoria) }}">\n <i class="bi bi-download me-1"></i>Baixar PDF\n </a>\n </div>\n </div>\n {% endfor %}\n </div>\n</section>\n{% endfor %}\n{% endblock %}'

TMPLS['admin.html'] = ('{% extends "base.html" %}\n{% block title %}Admin - Clube do Saber{% endblock %}\n'
'{% block head %}<style>.panel{background:#fff;border-radius:12px;padding:1.5rem;box-shadow:0 2px 8px rgba(0,0,0,.07);margin-bottom:1.5rem}</style>{% endblock %}\n'
'{% block content %}\n'
'<h4 class="fw-bold mb-4" style="color:#1a1a2e"><i class="bi bi-gear-fill me-2" style="color:#6C3483"></i>Painel Admin</h4>\n'
'{% if not mail_ok %}<div class="alert alert-warning"><i class="bi bi-envelope-x-fill me-2"></i>E-mail nao configurado. Defina <code>MAIL_USERNAME</code> e <code>MAIL_PASSWORD</code> no Render.</div>{% endif %}\n'
'{% if msg %}<div class="alert alert-{{ msg_type }} alert-dismissible fade show">{{ msg }}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>{% endif %}\n'
'<div class="panel">\n'
' <h6 class="fw-bold mb-3"><i class="bi bi-person-plus-fill me-2"></i>Adicionar Usuario</h6>\n'
' <form method="post" action="{{ url_for(\'admin_add\') }}" class="row g-2">\n'
'  <div class="col-md-3"><input name="username" class="form-control form-control-sm" placeholder="Usuario" required></div>\n'
'  <div class="col-md-3"><input name="email" type="email" class="form-control form-control-sm" placeholder="E-mail (opcional)"></div>\n'
'  <div class="col-md-2"><input name="password" class="form-control form-control-sm" placeholder="Senha" required></div>\n'
'  <div class="col-md-2"><select name="is_admin" class="form-select form-select-sm"><option value="0">Usuario</option><option value="1">Admin</option></select></div>\n'
'  <div class="col-md-2"><button class="btn btn-sm btn-brand w-100">Adicionar</button></div>\n'
' </form>\n'
'</div>\n'
'<div class="panel">\n'
' <h6 class="fw-bold mb-3"><i class="bi bi-file-earmark-excel-fill me-2 text-success"></i>Importar via Excel (.xlsx)</h6>\n'
' <p class="text-muted small mb-2">Colunas: <code>username</code>, <code>password</code>, <code>email</code> (opcional), <code>is_admin</code> (opcional).</p>\n'
' {% if not has_openpyxl %}<div class="alert alert-warning py-2">Instale: <code>pip install openpyxl</code></div>\n'
' {% else %}\n'
' <form method="post" action="{{ url_for(\'admin_import\') }}" enctype="multipart/form-data" class="d-flex gap-2 align-items-center flex-wrap">\n'
'  <input type="file" name="file" accept=".xlsx" class="form-control form-control-sm" style="max-width:280px" required>\n'
'  <button class="btn btn-sm btn-success">Importar</button>\n'
' </form>\n'
' {% if created is not none %}<div class="mt-2 small text-muted">Criados: <strong>{{ created }}</strong> | Pulados: <strong>{{ skipped }}</strong>{% if emailed %} | E-mails: <strong>{{ emailed }}</strong>{% endif %}</div>{% endif %}\n'
' {% endif %}\n'
'</div>\n'
'<div class="panel">\n'
' <h6 class="fw-bold mb-3"><i class="bi bi-people-fill me-2"></i>Usuarios ({{ users|length }})</h6>\n'
' <div class="table-responsive">\n'
'  <table class="table table-hover table-sm align-middle">\n'
'   <thead class="table-light"><tr><th>ID</th><th>Usuario</th><th>E-mail</th><th>Perfil</th><th>Categorias</th><th>Criado</th><th></th></tr></thead>\n'
'   <tbody>\n'
'   {% for u in users %}\n'
'   <tr>\n'
'    <td class="text-muted small">{{ u.id }}</td>\n'
'    <td class="fw-semibold">{{ u.username }}</td>\n'
'    <td class="text-muted small">{{ u.email or \'-\' }}</td>\n'
'    <td>{% if u.is_admin %}<span class="badge bg-warning text-dark">Admin</span>{% else %}<span class="badge bg-secondary">Usuario</span>{% endif %}</td>\n'
'    <td>\n'
'     {% if u.is_admin %}<span class="badge bg-success">Todas</span>\n'
'     {% elif u.categories %}<span class="badge bg-info text-dark">Restrito</span>\n'
'     {% else %}<span class="badge bg-success">Todas</span>{% endif %}\n'
'     {% if not u.is_admin %}<a href="{{ url_for(\'admin_user_cats\', uid=u.id) }}" class="btn btn-xs btn-outline-primary ms-1" style="font-size:.7rem;padding:1px 6px"><i class="bi bi-sliders"></i></a>{% endif %}\n'
'    </td>\n'
'    <td class="text-muted small">{{ (u.created_at or \'\')[:10] }}</td>\n'
'    <td>\n'
'     <form method="post" action="{{ url_for(\'admin_delete\', uid=u.id) }}" class="d-inline" onsubmit="return confirm(\'Deletar {{ u.username }}?\')">\n'
'      <button class="btn btn-sm btn-outline-danger py-0"><i class="bi bi-trash"></i></button>\n'
'     </form>\n'
'    </td>\n'
'   </tr>\n'
'   {% endfor %}\n'
'   </tbody>\n'
'  </table>\n'
' </div>\n'
'</div>\n'
'{% endblock %}')

TMPLS['user_cats.html'] = ('<!doctype html><html lang="pt-BR"><head>\n'
'<meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">\n'
'<title>Categorias — {{ user.username }}</title>\n'
'<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css">\n'
'<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.min.css">\n'
'<style>body{background:linear-gradient(135deg,#16213e,#533483);min-height:100vh;display:flex;align-items:center;justify-content:center}'
'.card{border:none;border-radius:16px;box-shadow:0 20px 60px rgba(0,0,0,.4)}'
'.form-check-input:checked{background-color:#6C3483;border-color:#6C3483}</style>\n'
'</head><body>\n'
'<div class="card p-4" style="width:100%;max-width:520px">\n'
' <div class="d-flex align-items-center gap-2 mb-3">\n'
'  <div style="width:40px;height:40px;background:#6C3483;border-radius:10px;display:flex;align-items:center;justify-content:center">'
'<i class="bi bi-sliders text-white"></i></div>\n'
'  <div><h5 class="fw-bold mb-0">Categorias de acesso</h5>'
'<p class="text-muted small mb-0">Usuário: <strong>{{ user.username }}</strong></p></div>\n'
' </div>\n'
' <form method="post">\n'
'  <div class="mb-1">\n'
'   <div class="form-check p-3 mb-2 rounded" style="background:#f8f5ff;border:2px solid #6C3483">\n'
'    <input class="form-check-input" type="checkbox" id="all_access" name="all_access" value="1"\n'
'     {% if not current %}checked{% endif %}\n'
'     onchange="document.querySelectorAll(\'.cat-chk\').forEach(c=>c.disabled=this.checked)">\n'
'    <label class="form-check-label fw-bold" for="all_access">\n'
'     <i class="bi bi-unlock-fill me-1 text-success"></i> Acesso a TODAS as categorias\n'
'    </label>\n'
'   </div>\n'
'   <p class="text-muted small mb-2">Ou selecione categorias específicas abaixo:</p>\n'
'   {% for cat in all_cats %}\n'
'   <div class="form-check py-1 px-3">\n'
'    <input class="form-check-input cat-chk" type="checkbox" name="cats" value="{{ cat }}" id="c{{ loop.index }}"\n'
'     {% if current and cat in current %}checked{% endif %}\n'
'     {% if not current %}disabled{% endif %}>\n'
'    <label class="form-check-label" for="c{{ loop.index }}">{{ cat }}</label>\n'
'   </div>\n'
'   {% endfor %}\n'
'  </div>\n'
'  <div class="d-flex gap-2 mt-3">\n'
'   <button class="btn text-white fw-bold" style="background:#6C3483"><i class="bi bi-check-lg me-1"></i>Salvar</button>\n'
'   <a href="{{ url_for(\'admin\') }}" class="btn btn-outline-secondary">Cancelar</a>\n'
'  </div>\n'
' </form>\n'
'</div>\n'
'</body></html>')

TMPLS['search.html'] = ('{% extends "base.html" %}\n{% block title %}Busca - Clube do Saber{% endblock %}\n{% block head %}<style>\n.book-card{background:#fff;border-radius:12px;padding:1.2rem 1.4rem;box-shadow:0 2px 8px rgba(0,0,0,.07);display:flex;flex-direction:column;height:100%;border-top:4px solid var(--cc);transition:box-shadow .15s}\n.book-card:hover{box-shadow:0 6px 18px rgba(0,0,0,.13)}\n.book-title{font-weight:700;font-size:.95rem;line-height:1.3;color:#1a1a2e;margin-bottom:.3rem}\n.book-sub{font-size:.8rem;color:#6c757d;line-height:1.3;flex:1}\n.btn-dl{font-size:.8rem;padding:.35rem .9rem;border-radius:8px;margin-top:.9rem}\n</style>{% endblock %}\n{% block content %}\n<h4 class="fw-bold mb-2" style="color:#1a1a2e"><i class="bi bi-search me-2" style="color:#6C3483"></i>Pesquisar</h4>\n<form action="{{ url_for("search") }}" method="get" class="mb-4">\n <div class="input-group" style="max-width:480px">\n <input type="search" name="q" class="form-control" placeholder="Titulo, categoria, subcategoria..." value="{{ q }}" autofocus>\n <button class="btn" style="background:#6C3483;color:#fff" type="submit"><i class="bi bi-search me-1"></i>Buscar</button>\n </div>\n</form>\n{% if q %}\n {% if results %}\n <p class="text-muted small mb-3">{{ results|length }} resultado{% if results|length != 1 %}s{% endif %} para <strong>"{{ q }}"</strong></p>\n <div class="row g-3">\n {% for book in results %}\n <div class="col-6 col-md-4 col-lg-3">\n <div class="book-card" style="--cc:{{ cat_color(book.categoria) }}">\n <span class="badge mb-2 fw-normal" style="background:{{ cat_color(book.categoria) }};font-size:.7rem">{{ book.categoria }}</span>\n <div class="book-title">{{ book.titulo }}</div>\n <div class="book-sub">{{ book.subtitulo }}</div>\n <a href="{{ url_for("download", filename=book.filename) }}" class="btn btn-sm btn-dl text-white" style="background:{{ cat_color(book.categoria) }}">\n <i class="bi bi-download me-1"></i>Baixar PDF\n </a>\n </div>\n </div>\n {% endfor %}\n </div>\n {% else %}\n <div class="text-center py-5 text-muted">\n <i class="bi bi-search" style="font-size:3rem;opacity:.3"></i>\n <p class="mt-3">Nenhum resultado para <strong>"{{ q }}"</strong>.</p>\n <a href="{{ url_for("dashboard") }}" class="btn btn-sm btn-outline-secondary">Ver todos os livros</a>\n </div>\n {% endif %}\n{% endif %}\n{% endblock %}')

app = Flask(__name__, static_folder=None)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.jinja_env.loader = DictLoader(TMPLS)
app.jinja_env.globals.update(categories=CATEGORIES, cat_icon=_cat_icon, cat_color=_cat_color)

def get_db():
    db = sqlite3.connect(DB_PATH)
    db.row_factory = sqlite3.Row
    return db

def init_db():
    db = get_db()
    sql = ('CREATE TABLE IF NOT EXISTS users '
           '(id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, '
           'email TEXT DEFAULT \'\', password TEXT NOT NULL, is_admin INTEGER DEFAULT 0, '
           'categories TEXT DEFAULT NULL, '
           'created_at TEXT DEFAULT (datetime(\'now\'))); '
           'CREATE TABLE IF NOT EXISTS reset_tokens '
           '(id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, '
           'token TEXT UNIQUE NOT NULL, expires_at TEXT NOT NULL, used INTEGER DEFAULT 0);')
    db.executescript(sql)
    # migration: adiciona coluna categories se ainda não existir
    try:
        db.execute('ALTER TABLE users ADD COLUMN categories TEXT DEFAULT NULL')
        db.commit()
    except Exception:
        pass
    if not db.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        db.execute('INSERT INTO users (username, password, is_admin) VALUES (?,?,1)',
                   ('admin', generate_password_hash('admin123')))
    db.commit(); db.close()

init_db()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'): abort(403)
        return f(*args, **kwargs)
    return decorated

def _admin_ctx(**extra):
    db = get_db()
    users = db.execute('SELECT * FROM users ORDER BY id').fetchall()
    db.close()
    ctx = dict(users=users, mail_ok=bool(MAIL_USER and MAIL_PASS),
               has_openpyxl=HAS_OPENPYXL, msg=None, msg_type='success',
               created=None, skipped=None, emailed=None)
    ctx.update(extra)
    return ctx

def _get_user_categories():
    """Retorna set de categorias permitidas ou None (= todas)."""
    if session.get('is_admin'):
        return None
    cats_json = session.get('user_cats')
    if cats_json:
        return set(json.loads(cats_json))
    return None  # sem restrição

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session: return redirect(url_for('dashboard'))
    error = None
    if request.method == 'POST':
        un = request.form['username'].strip(); pw = request.form['password']
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE username=?', (un,)).fetchone()
        db.close()
        if user and check_password_hash(user['password'], pw):
            session.permanent = True
            session['user_id'] = user['id']; session['username'] = user['username']
            session['is_admin'] = bool(user['is_admin'])
            session['user_cats'] = user['categories']  # JSON string ou None
            return redirect(url_for('dashboard'))
        error = 'Usuario ou senha incorretos.'
    return render_template_string(TMPLS['login.html'], error=error)

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/forgot', methods=['GET', 'POST'])
def forgot():
    msg = None; msg_type = 'info'
    if request.method == 'POST':
        email = request.form['email'].strip()
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE email=?', (email,)).fetchone()
        if user:
            token = secrets.token_urlsafe(32)
            expires = (datetime.utcnow() + timedelta(hours=1)).isoformat()
            db.execute('INSERT INTO reset_tokens (user_id, token, expires_at) VALUES (?,?,?)',
                       (user['id'], token, expires))
            db.commit()
            reset_url = url_for('reset_pw', token=token, _external=True)
            sent = _send_email(email, 'Redefinir senha - Clube do Saber',
                               _reset_html(user['username'], reset_url))
            msg = 'Link enviado!' if sent else 'E-mail nao configurado. Contate o admin.'
            msg_type = 'success' if sent else 'warning'
        else:
            msg = 'Se esse e-mail estiver cadastrado, o link sera enviado.'
        db.close()
    return render_template_string(TMPLS['forgot.html'], msg=msg, msg_type=msg_type)

@app.route('/reset/<token>', methods=['GET', 'POST'])
def reset_pw(token):
    db = get_db()
    row = db.execute('SELECT * FROM reset_tokens WHERE token=? AND used=0', (token,)).fetchone()
    if not row or datetime.utcnow().isoformat() > row['expires_at']:
        db.close()
        return render_template_string(TMPLS['reset.html'], expired=True, error=None)
    error = None
    if request.method == 'POST':
        pw = request.form['password']; pw2 = request.form['password2']
        if pw != pw2: error = 'As senhas nao coincidem.'
        elif len(pw) < 6: error = 'Minimo 6 caracteres.'
        else:
            db.execute('UPDATE users SET password=? WHERE id=?',
                       (generate_password_hash(pw), row['user_id']))
            db.execute('UPDATE reset_tokens SET used=1 WHERE id=?', (row['id'],))
            db.commit(); db.close()
            return redirect(url_for('login'))
    db.close()
    return render_template_string(TMPLS['reset.html'], expired=False, error=error)

@app.route('/dashboard')
@login_required
def dashboard():
    allowed = _get_user_categories()
    if allowed is not None:
        filtered = {k: v for k, v in CATEGORIES.items() if k in allowed}
    else:
        filtered = CATEGORIES
    return render_template_string(TMPLS['dashboard.html'], categories=filtered)

@app.route('/download/<path:filename>')
@login_required
def download(filename):
    import re
    safe = os.path.basename(filename)
    # Verifica acesso à categoria do livro
    allowed = _get_user_categories()
    book_meta = next((b for b in EBOOKS if b['filename'] == safe), None)
    if allowed is not None:
        if book_meta and book_meta['categoria'] not in allowed:
            abort(403)
    # Se o ebook tem driveId, redirecionar para Google Drive
    if book_meta and book_meta.get('driveId'):
        return redirect('https://drive.google.com/uc?export=download&id={}'.format(book_meta['driveId']))
    for folder in [EBOOKS_DIR, os.path.join(EBOOKS_DIR, 'livros_para_upload')]:
        path = os.path.join(folder, safe)
        if os.path.exists(path):
            clean_name = re.sub(r'^\d+_', '', safe)
            return send_from_directory(folder, safe, as_attachment=True,
                                       download_name=clean_name)
    abort(404)

@app.route('/admin')
@login_required
@admin_required
def admin():
    return render_template_string(TMPLS['admin.html'], **_admin_ctx())

@app.route('/admin/users/<int:uid>/categories', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_user_cats(uid):
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE id=?', (uid,)).fetchone()
    if not user:
        db.close(); abort(404)
    if request.method == 'POST':
        all_access = request.form.get('all_access')
        if all_access:
            cats_json = None  # acesso total
        else:
            selected = request.form.getlist('cats')
            cats_json = json.dumps(selected) if selected else None
        db.execute('UPDATE users SET categories=? WHERE id=?', (cats_json, uid))
        db.commit(); db.close()
        return redirect(url_for('admin'))
    current = json.loads(user['categories']) if user['categories'] else None
    db.close()
    return render_template_string(TMPLS['user_cats.html'],
                                   user=user,
                                   all_cats=list(CATEGORIES.keys()),
                                   current=current)

@app.route('/admin/users/add', methods=['POST'])
@login_required
@admin_required
def admin_add():
    un = request.form['username'].strip(); em = request.form.get('email','').strip()
    pw = request.form['password']; ia = int(request.form.get('is_admin',0))
    db = get_db()
    try:
        db.execute('INSERT INTO users (username, email, password, is_admin) VALUES (?,?,?,?)',
                   (un, em, generate_password_hash(pw), ia))
        db.commit()
        sent = _send_email(em, 'Bem-vindo ao Clube do Saber!',
                           _welcome_html(un, pw, url_for('login', _external=True))) if em else False
        db.close()
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Usuario {} criado{}.'.format(un,' e e-mail enviado' if sent else ''),
                                                   msg_type='success'))
    except sqlite3.IntegrityError:
        db.close()
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Usuario ja existe.', msg_type='danger'))

@app.route('/admin/users/delete/<int:uid>', methods=['POST'])
@login_required
@admin_required
def admin_delete(uid):
    if uid == session['user_id']:
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Voce nao pode se deletar.', msg_type='danger'))
    db = get_db()
    db.execute('DELETE FROM users WHERE id=?', (uid,))
    db.commit(); db.close()
    return render_template_string(TMPLS['admin.html'],
                                  **_admin_ctx(msg='Usuario removido.', msg_type='success'))

@app.route('/admin/import', methods=['POST'])
@login_required
@admin_required
def admin_import():
    if not HAS_OPENPYXL:
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='openpyxl nao instalado.', msg_type='danger'))
    f = request.files.get('file')
    if not f or not f.filename.endswith('.xlsx'):
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Envie um arquivo .xlsx.', msg_type='danger'))
    created = skipped = emailed = 0
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
        def col(row, name):
            try:
                i = headers.index(name); v = row[i].value
                return str(v).strip() if v is not None else ''
            except ValueError: return ''
        db = get_db(); login_url = url_for('login', _external=True)
        for row in ws.iter_rows(min_row=2):
            un = col(row,'username'); pw = col(row,'password')
            if not un or not pw: continue
            em = col(row,'email')
            ia = 1 if col(row,'is_admin') in ('1','true','sim','yes','admin') else 0
            try:
                db.execute('INSERT INTO users (username, email, password, is_admin) VALUES (?,?,?,?)',
                           (un, em, generate_password_hash(pw), ia))
                db.commit(); created += 1
                if em and _send_email(em,'Bem-vindo ao Clube do Saber!',_welcome_html(un,pw,login_url)): emailed += 1
            except sqlite3.IntegrityError: skipped += 1
        db.close()
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Importacao: {} criados, {} pulados.'.format(created,skipped),
                                                   msg_type='success', created=created, skipped=skipped, emailed=emailed))
    except Exception as e:
        return render_template_string(TMPLS['admin.html'],
                                      **_admin_ctx(msg='Erro: {}'.format(e), msg_type='danger'))

@app.route('/search')
@login_required
def search():
    q = request.args.get('q', '').strip()
    results = []
    if q:
        ql = q.lower()
        allowed = _get_user_categories()
        for book in EBOOKS:
            if allowed is not None and book.get('categoria') not in allowed:
                continue
            haystack = ' '.join([
                book.get('titulo', ''),
                book.get('subtitulo', ''),
                book.get('categoria', ''),
                book.get('subcategoria', ''),
            ]).lower()
            if ql in haystack:
                results.append(book)
    return render_template_string(TMPLS['search.html'], q=q, results=results)

if __name__ == '__main__':
    app.run(debug=True)
